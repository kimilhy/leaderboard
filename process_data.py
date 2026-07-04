"""
Drone Racing Leaderboard Processor
Reads XLSX, removes false laps, classifies by pilot & try, computes stats, outputs JSON.
"""
import json
import statistics
from collections import defaultdict
import openpyxl


def parse_time(val):
    """Parse time string with European decimal comma to float seconds."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    return float(str(val).replace(",", ".").strip())


def load_data(filepath):
    """Load and parse XLSX into list of lap dicts."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Sheet1"]

    laps = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        laps.append({
            "round": int(row[0]),
            "race": int(row[1]),
            "race_start": str(row[2]) if row[2] else "",
            "pilot": str(row[3]).strip(),
            "lap_number": int(row[4]),
            "length": parse_time(row[5]),
            "race_time": parse_time(row[6]),
        })
    return laps


def remove_outliers_iqr(laps, multiplier=2.0):
    """Remove outlier laps using IQR method on lap times."""
    lengths = [l["length"] for l in laps]
    if len(lengths) < 4:
        return laps

    sorted_l = sorted(lengths)
    q1 = statistics.median(sorted_l[: len(sorted_l) // 2])
    q3 = statistics.median(sorted_l[(len(sorted_l) + 1) // 2 :])
    iqr = q3 - q1
    lower = q1 - multiplier * iqr
    upper = q3 + multiplier * iqr

    return [l for l in laps if lower <= l["length"] <= upper]


def best_consecutive_in_try(laps, count=3):
    """
    Find best N consecutive laps within a single try.
    Laps are sorted by lap_number — they're truly consecutive within a try.
    """
    sorted_laps = sorted(laps, key=lambda l: l["lap_number"])
    if len(sorted_laps) < count:
        return None

    best_sum = float("inf")
    best_window = None

    for i in range(len(sorted_laps) - count + 1):
        window = sorted_laps[i:i + count]
        # Verify they're actually consecutive lap numbers
        nums = [l["lap_number"] for l in window]
        if nums != list(range(min(nums), max(nums) + 1)):
            continue
        window_sum = sum(l["length"] for l in window)
        if window_sum < best_sum:
            best_sum = window_sum
            best_window = window

    if best_window is None:
        return None

    return {
        "sum": round(best_sum, 3),
        "avg": round(best_sum / count, 3),
        "laps": [{"lap": l["lap_number"], "time": round(l["length"], 3)} for l in best_window],
    }


def process(filepath, min_lap_time=10.0, max_lap_time=70.0):
    """
    Main processing pipeline.

    Filtering is purely physics-based — the XLSX "Valid" column is ignored.
    min_lap_time: laps faster than this are physically impossible (sensor errors)
    max_lap_time: laps slower than this are crashes/recoveries
    """
    all_laps = load_data(filepath)

    # Step 1: Remove physically impossible laps (hard min/max cutoffs)
    filtered = [l for l in all_laps if min_lap_time <= l["length"] <= max_lap_time]
    hard_removed = len(all_laps) - len(filtered)
    if hard_removed:
        print(f"Hard min/max filter ({min_lap_time}s–{max_lap_time}s): removed {hard_removed} physically impossible laps")

    # Step 2: Group into tries: (pilot, round, race)
    tries = defaultdict(list)
    for lap in filtered:
        key = (lap["pilot"], lap["round"], lap["race"])
        tries[key].append(lap)

    # Step 3: Deduplicate same lap number within a try (keep shortest time)
    deduped_tries = {}
    dupe_count = 0
    for key, try_laps in tries.items():
        best = {}
        for l in try_laps:
            if l["lap_number"] not in best or l["length"] < best[l["lap_number"]]["length"]:
                best[l["lap_number"]] = l
        dupe_count += len(try_laps) - len(best)
        deduped_tries[key] = list(best.values())
    if dupe_count:
        print(f"Dedup duplicate lap numbers: removed {dupe_count}")

    # Step 4: Remove statistical outliers per try
    cleaned_tries = {}
    for key, try_laps in deduped_tries.items():
        cleaned_tries[key] = remove_outliers_iqr(try_laps)

    total_before_iqr = sum(len(v) for v in deduped_tries.values())
    total_after_iqr = sum(len(v) for v in cleaned_tries.values())
    iqr_removed = total_before_iqr - total_after_iqr
    if iqr_removed:
        print(f"IQR outlier filter: removed {iqr_removed} statistical outliers")

    # Step 5: Build per-try stats + per-pilot aggregates
    pilot_tries = defaultdict(list)
    for (pilot, rnd, race), try_laps in sorted(cleaned_tries.items()):
        try_laps_sorted = sorted(try_laps, key=lambda l: l["lap_number"])
        lap_times = [l["length"] for l in try_laps_sorted]
        b3 = best_consecutive_in_try(try_laps_sorted)

        pilot_tries[pilot].append({
            "round": rnd,
            "race": race,
            "race_start": try_laps_sorted[0]["race_start"],
            "lap_count": len(try_laps_sorted),
            "best_lap": round(min(lap_times), 3),
            "avg_lap": round(statistics.mean(lap_times), 3),
            "best_consecutive_3": b3,
            "laps": [{"lap": l["lap_number"], "time": round(l["length"], 3)} for l in try_laps_sorted],
        })

    # Step 6: Build per-pilot summary
    results = []
    for pilot, pilot_try_list in sorted(pilot_tries.items()):
        all_times = []
        for t in pilot_try_list:
            all_times.extend(l["time"] for l in t["laps"])

        all_times.sort()
        best_lap = min(all_times)
        avg_lap = round(statistics.mean(all_times), 3)
        median_lap = round(statistics.median(all_times), 3)
        total_laps = len(all_times)

        # Best 3 consecutive across all tries (pick the try with best b3 avg)
        best_b3 = None
        best_b3_try = None
        for t in pilot_try_list:
            if t["best_consecutive_3"]:
                if best_b3 is None or t["best_consecutive_3"]["avg"] < best_b3["avg"]:
                    best_b3 = t["best_consecutive_3"]
                    best_b3_try = f"Round {t['round']} Race {t['race']}"

        results.append({
            "pilot": pilot,
            "total_laps": total_laps,
            "total_tries": len(pilot_try_list),
            "best_lap": best_lap,
            "avg_lap": avg_lap,
            "median_lap": median_lap,
            "best_consecutive_3": best_b3,
            "best_consecutive_3_try": best_b3_try,
            "tries": pilot_try_list,
        })

    # Sort by best 3 consecutive avg (ascending → faster = better), fallback to avg lap
    results.sort(key=lambda r: (
        r["best_consecutive_3"]["avg"] if r["best_consecutive_3"] else float("inf"),
        r["avg_lap"],
    ))

    output = {
        "source": filepath,
        "config": {"min_lap_time_s": min_lap_time, "max_lap_time_s": max_lap_time},
        "total_laps_raw": len(all_laps),
        "total_laps_hard_filter_removed": hard_removed,
        "total_laps_dup_removed": dupe_count,
        "total_laps_iqr_removed": iqr_removed,
        "total_laps_final": sum(r["total_laps"] for r in results),
        "pilots": len(results),
        "results": results,
    }
    return output


if __name__ == "__main__":
    import sys, glob
    # Accept filename as argument, or auto-detect the only .xlsx in the directory
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
    else:
        xlsx_files = glob.glob("*.xlsx")
        if not xlsx_files:
            sys.exit("No .xlsx file found. Drop one in the directory or pass it as an argument.")
        if len(xlsx_files) > 1:
            sys.exit(f"Multiple .xlsx files found: {xlsx_files}. Pass the one to use as an argument.")
        filepath = xlsx_files[0]
    data = process(filepath)
    with open("data.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Processed {data['total_laps_raw']} total laps → {data['total_laps_final']} clean laps")
    print(f"  Hard-filter: {data['total_laps_hard_filter_removed']} | Dupes: {data['total_laps_dup_removed']} | IQR: {data['total_laps_iqr_removed']}")
    print(f"Pilots: {data['pilots']}")
    for r in data["results"]:
        b3 = r["best_consecutive_3"]
        print(f"  {r['pilot']}: {r['total_laps']} laps in {r['total_tries']} tries | Best: {r['best_lap']}s | Avg: {r['avg_lap']}s | Best3: {b3['avg'] if b3 else 'N/A'}s ({r.get('best_consecutive_3_try', '')})")
